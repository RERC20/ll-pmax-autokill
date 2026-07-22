#!/usr/bin/env python3
# kill_engine_google_auto.py
# ---------------------------------------------------------------------------
# UNATTENDED Google-direct auto-kill — for running on a schedule (e.g. Claude
# schedule / GitHub repo, hourly). Runs once when CALLED, no user input:
#   1. Pull the live feed (Google Ads cost/clicks + Shopify revenue, UK tz).
#   2. Tag new winners (any Shopify sale) + fast-path them into the Winners campaign.
#   3. WINNER PACE RULE (v11): kill winners whose Winners-campaign spend since
#      their last sale exceeds that sale's revenue / 2.3 (see block below).
#   4. Apply the SAME v4 rules (evaluate) to the TESTING pool (winners exempt).
#   5. DRAFT every flagged product in Shopify (no yes/no prompt).
#   6. NOTIFY:
#        - TELEGRAM every run  -> run stats + the .xlsx (instant push, no daily cap)
#        - RESEND email twice/day (SUMMARY_HOURS) -> a TEXT digest of the last 12h kills + the .xlsx
#
# Separate from kill_engine_google.py (which asks for confirmation) — it changes
# nothing in the other files; it only REUSES their tested functions.
#
# Usage:
#   python kill_engine_google_auto.py          # live: drafts the kills + notifies
#   python kill_engine_google_auto.py --dry     # safe test: computes + notifies, DRAFTS NOTHING
#   python kill_engine_google_auto.py --test    # like the run but also FORCES the 12h email now (testing)
#
# Setup (env vars, or the git-ignored _secrets_local.py):
#   TELEGRAM_BOT_TOKEN + TELEGRAM_CHAT_ID  (from @BotFather; drives the every-run push)
#   RESEND_API_KEY  (free key from resend.com; twice-daily digest; signup inbox == EMAIL_TO)
# Any unset channel is simply skipped (the run still drafts + logs).
# ---------------------------------------------------------------------------
import sys, os, csv, base64, html, datetime, collections, requests
from openpyxl import Workbook
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from zoneinfo import ZoneInfo

# reuse the EXACT feed + rules + Shopify write + logging from the existing engines
from kill_engine_google import build_feed, PMAX_CAMPAIGN_ID as WINNERS_CAMPAIGN_ID
from kill_engine_v4 import evaluate, shopify_token, shopify_draft, _Tee, SHOP, SHOP_API
from creds import cred                          # env -> _secrets_local.py -> '' (so local testing works too)

UK = ZoneInfo('Europe/London')                 # store + Google Ads account run on UK time

# ---- notifications ----
EMAIL_TO       = cred('EMAIL_TO')               # digest inbox — from Secrets/_secrets_local, never in public code
RESEND_API_KEY = cred('RESEND_API_KEY')
RESEND_FROM    = cred('RESEND_FROM') or 'onboarding@resend.dev'
TELEGRAM_TOKEN = cred('TELEGRAM_BOT_TOKEN')     # from @BotFather — every-run push
TELEGRAM_CHAT  = cred('TELEGRAM_CHAT_ID')       # your numeric chat id (@userinfobot)
SUMMARY_HOURS  = (9, 21)                        # UK hours for the twice-a-day (every 12h) email digest
RUN_LOG        = 'kill_engine_auto_runs.log'
KILLS_LOG      = 'kills_log_auto.csv'

# ── Telegram kill formatting: plain-English tier + the rule's reason + core metrics ──
TIER_LABEL = {
    'Tier 1': 'Tier 1 — no sale (70+ clicks)', 'Tier 2': 'Tier 2 — no sale (£5+ or 40+ clicks)',
    'Tier 3': 'Tier 3 — Mon stale no-sale',    'Tier 4': 'Tier 4 — Mon ghost (<5 clicks)',
    'Tier 5': 'Tier 5 — stalled winner',       'Tier 6': 'Tier 6 — below 2.0 ROAS (7d)',
    'Tier 7': 'Tier 7 — slow dribbler (30d)'}
def _fmt_kill(p, tier, why, run_date):
    dl = (run_date - datetime.date.fromisoformat(p['pub'])).days
    return (f"• <b>{html.escape(p['name'][:46])}</b>\n"
            f"  <code>{p['pid']}</code> · <b>{TIER_LABEL.get(tier, tier)}</b> · {dl}d live\n"
            f"  ↳ why: {html.escape(str(why))}\n"
            f"  ↳ spend 7d £{p['cost7']:.2f} / 30d £{p['cost30']:.2f} · ROAS7 {p['roas7']:.2f} · "
            f"rev 30d £{p['rev30']:.2f} · {p['clicks30']} clk")

def _days_live(p, run_date):
    return (run_date - datetime.date.fromisoformat(p['pub'])).days

# ── WINNERS (w_campaign): tag + exempt (owner-approved 2026-07-11) ──────────
# A product with >=1 Shopify sale (GROUND TRUTH — never depends on Google's
# under-reporting) that is still ACTIVE is a "winner". It gets the w_campaign
# tag, which does two things:
#   (a) Simprosys rule maps tag -> custom_label_1=w_campaign -> the product hops
#       from the Testing PMax campaign to the Winners campaign on the next
#       feed sync (near-real-time, tag changes fire Shopify webhooks);
#   (b) EXEMPTS it from this engine's kill rules — winners will get their OWN
#       rules later; until then NO product with a sale is ever drafted here.
#       Only never-sold products (Tiers 1-4 territory) keep dying.
# First sales always appear in rev30 (live Shopify orders pull) within one
# 8-min run. Sales older than 30d were tagged by the one-off backfill
# (backfill_winner_tags.py, run 2026-07-11) — so rev30>0 is a complete signal
# for every NEW first sale going forward.
WINNER_TAG = 'w_campaign'
LOST_TAG   = 'l_camp'      # "lost" — WAS a winner, killed by the pace rule (v11). Set on winner
                           # kills; stripped again if the product is ever resurrected and re-sells.

def shopify_add_tag(tok, pid, tag):
    m = 'mutation($id:ID!,$t:[String!]!){tagsAdd(id:$id,tags:$t){userErrors{message}}}'
    try:
        j = requests.post(f"https://{SHOP}/admin/api/{SHOP_API}/graphql.json",
                          headers={'X-Shopify-Access-Token': tok, 'Content-Type': 'application/json'},
                          json={'query': m, 'variables': {'id': f"gid://shopify/Product/{pid}", 't': [tag]}},
                          timeout=30).json()
        errs = (j.get('data', {}).get('tagsAdd') or {}).get('userErrors') or []
        return 'ok' if not errs else f"err: {errs[0].get('message', '?')[:60]}"
    except Exception as ex:
        return f"err: {ex}"

def shopify_remove_tag(tok, pid, tag):
    m = 'mutation($id:ID!,$t:[String!]!){tagsRemove(id:$id,tags:$t){userErrors{message}}}'
    try:
        j = requests.post(f"https://{SHOP}/admin/api/{SHOP_API}/graphql.json",
                          headers={'X-Shopify-Access-Token': tok, 'Content-Type': 'application/json'},
                          json={'query': m, 'variables': {'id': f"gid://shopify/Product/{pid}", 't': [tag]}},
                          timeout=30).json()
        errs = (j.get('data', {}).get('tagsRemove') or {}).get('userErrors') or []
        return 'ok' if not errs else f"err: {errs[0].get('message', '?')[:60]}"
    except Exception as ex:
        return f"err: {ex}"

def shopify_set_label_metafield(tok, pid, value=WINNER_TAG):
    """Also write Simprosys's own attribute metafield (mm-google-shopping.custom_label_1).
    The app's bulk-edit stores labels app-side, but it READS this metafield on its syncs —
    so future winners pick up the feed label without a human touching the app. The Ads-API
    item-ID mover (added after the campaigns exist) is the guaranteed instant path.
    value: WINNER_TAG (winners) or CHAMPION_TAG (champions tier, 2026-07-20)."""
    m = ('mutation($mf:[MetafieldsSetInput!]!){metafieldsSet(metafields:$mf){userErrors{message}}}')
    v = {'mf': [{'ownerId': f"gid://shopify/Product/{pid}", 'namespace': 'mm-google-shopping',
                 'key': 'custom_label_1', 'type': 'single_line_text_field', 'value': value}]}
    try:
        j = requests.post(f"https://{SHOP}/admin/api/{SHOP_API}/graphql.json",
                          headers={'X-Shopify-Access-Token': tok, 'Content-Type': 'application/json'},
                          json={'query': m, 'variables': v}, timeout=30).json()
        errs = (j.get('data', {}).get('metafieldsSet') or {}).get('userErrors') or []
        return 'ok' if not errs else f"err: {errs[0].get('message', '?')[:60]}"
    except Exception as ex:
        return f"err: {ex}"

def tag_new_winners(feed, dry):
    """Tag every ACTIVE product with a Shopify sale (rev30>0) not yet tagged w_campaign."""
    new = [p for p in feed if p['rev30'] > 0 and WINNER_TAG not in p['tags']]
    if not new:
        return []
    tok = None if dry else shopify_token()
    for p in new:
        res = 'DRY (not tagged)' if dry else shopify_add_tag(tok, p['pid'], WINNER_TAG)
        mres = 'DRY' if dry else shopify_set_label_metafield(tok, p['pid'])
        if LOST_TAG in p['tags'] and not dry:
            shopify_remove_tag(tok, p['pid'], LOST_TAG)   # resurrected + sold again: no longer "lost"
        print(f"  {'would tag' if dry else 'tag'} winner {p['pid']} -> {res} (label metafield: {mres}) | {p['name'][:42]}")
        p['tags'].append(WINNER_TAG)     # exempt from kill rules in THIS same run too
    return new

# ── ADS FAST-PATH: move new winners between the two PMax campaigns INSTANTLY ─
# The Testing/Winners split (2026-07-11) filters on custom_label_1=w_campaign,
# but that label rides Simprosys's feed sync (minutes-to-hours). This path skips
# the wait: the moment a product wins, its variant item-ids are written straight
# into the campaigns' listing trees via the Ads API —
#   Winners  (asset group 6684080392): item-ids UNIT_INCLUDED  (serves NOW)
#   Testing  (asset group 6729681029): item-ids UNIT_EXCLUDED  (stops NOW)
# Both trees' "everything else" branches were pre-converted to item-id
# subdivisions on 2026-07-11, so this is a plain node-add. Idempotent (checks
# existing nodes first). If it ever fails, the label path still moves the
# product on the next feed sync — so failures WARN, never break the kill run.
WINNERS_AG_ID = '6684080392'
TESTING_AG_ID = '6729681029'

def _ads_search(ga, gt, query):
    out = []; tok = None
    while True:
        body = {'query': query}
        if tok: body['pageToken'] = tok
        r = requests.post(f"{ga.ADS_BASE}/customers/{ga.CUSTOMER_ID}/googleAds:search",
                          headers=ga._headers(gt), json=body, timeout=60).json()
        if 'error' in r: raise RuntimeError(str(r)[:300])
        out += r.get('results', []); tok = r.get('nextPageToken')
        if not tok: return out

def _variant_item_ids(stok, pid):
    q = 'query($id:ID!){product(id:$id){variants(first:100){edges{node{legacyResourceId}}}}}'
    j = requests.post(f"https://{SHOP}/admin/api/{SHOP_API}/graphql.json",
                      headers={'X-Shopify-Access-Token': stok, 'Content-Type': 'application/json'},
                      json={'query': q, 'variables': {'id': f"gid://shopify/Product/{pid}"}}, timeout=30).json()
    vs = [e['node']['legacyResourceId'] for e in j['data']['product']['variants']['edges']]
    return [f"shopify_zz_{pid}_{v}".lower() for v in vs]

def ads_fast_path(new_winners, stok):
    try:
        import google_ads_connect as ga
        gt = ga.get_access_token()
        rows = _ads_search(ga, gt,
            "SELECT asset_group.id, asset_group_listing_group_filter.resource_name, "
            "asset_group_listing_group_filter.type, asset_group_listing_group_filter.parent_listing_group_filter, "
            "asset_group_listing_group_filter.case_value.product_custom_attribute.index, "
            "asset_group_listing_group_filter.case_value.product_custom_attribute.value, "
            "asset_group_listing_group_filter.case_value.product_item_id.value "
            "FROM asset_group_listing_group_filter WHERE asset_group.id IN (6684080392,6729681029)")
        info = {}
        for agid in (WINNERS_AG_ID, TESTING_AG_ID):
            ag = [x for x in rows if str(x['assetGroup']['id']) == agid]
            subdiv = None; have = set()
            for x in ag:   # the item-id subdivision = the SUBDIVISION whose case is attr1-with-no-value
                f = x['assetGroupListingGroupFilter']; cv = f.get('caseValue', {})
                pca = cv.get('productCustomAttribute')
                if f['type'] == 'SUBDIVISION' and pca is not None and 'value' not in pca:
                    subdiv = f['resourceName']
            for x in ag:
                f = x['assetGroupListingGroupFilter']; cv = f.get('caseValue', {})
                if f.get('parentListingGroupFilter') == subdiv and cv.get('productItemId', {}).get('value'):
                    have.add(cv['productItemId']['value'].lower())
            if not subdiv: raise RuntimeError(f"item-id subdivision not found in AG {agid}")
            info[agid] = (subdiv, have)
        made = 0
        for agid, node_type in ((WINNERS_AG_ID, 'UNIT_INCLUDED'), (TESTING_AG_ID, 'UNIT_EXCLUDED')):
            subdiv, have = info[agid]
            ops = []
            for p in new_winners:
                for iid in _variant_item_ids(stok, p['pid']):
                    if iid not in have:
                        ops.append({"create": {"assetGroup": f"customers/{ga.CUSTOMER_ID}/assetGroups/{agid}",
                                               "type": node_type, "listingSource": "SHOPPING",
                                               "parentListingGroupFilter": subdiv,
                                               "caseValue": {"productItemId": {"value": iid}}}})
            if ops:
                r = requests.post(f"{ga.ADS_BASE}/customers/{ga.CUSTOMER_ID}/assetGroupListingGroupFilters:mutate",
                                  headers=ga._headers(gt), json={"operations": ops}, timeout=60).json()
                if 'error' in r: raise RuntimeError(str(r)[:300])
                made += len(ops)
        print(f"  fast-path: {made} item-id nodes written (Winners include / Testing exclude)")
        return made, None
    except Exception as ex:
        print(f"  !! fast-path FAILED (harmless - label path still moves it on next feed sync): {ex}")
        return 0, str(ex)[:150]

# ── BEST SELLERS COLLECTION AUTO-ADD (owner request 2026-07-13) ─────────────
# Every ACTIVE winner (w_campaign) is ensured a member of the manual "Best
# Sellers" collection (handle best-sellers, shown on the storefront/product
# pages). ADD-ONLY by design: a winner later drafted/killed STAYS a member
# (owner call — drafts don't render on the storefront anyway, and a
# reactivated product is already in place). Collection sortOrder is
# BEST_SELLING, so Shopify auto-ranks the page by real sales. Idempotent
# every 8-min run; failures WARN, never break the run.
BESTSELLER_COLLECTION_ID = '690375426428'    # "Best Sellers" (manual collection)

def sync_bestseller_collection(tok, dry):
    """Add any active winner missing from the Best Sellers collection. Returns (added, err)."""
    try:
        cgid = f"gid://shopify/Collection/{BESTSELLER_COLLECTION_ID}"
        Q = ('query($id:ID!,$c:String){collection(id:$id){products(first:250,after:$c){'
             'pageInfo{hasNextPage endCursor} edges{node{legacyResourceId}}}}}')
        have = set(); cur = None
        while True:
            j = requests.post(f"https://{SHOP}/admin/api/{SHOP_API}/graphql.json",
                              headers={'X-Shopify-Access-Token': tok, 'Content-Type': 'application/json'},
                              json={'query': Q, 'variables': {'id': cgid, 'c': cur}}, timeout=60).json()
            co = (j.get('data') or {}).get('collection')
            if co is None: return 0, 'Best Sellers collection not found'
            pp = co['products']
            for e in pp['edges']: have.add(str(e['node']['legacyResourceId']))
            if not pp['pageInfo']['hasNextPage']: break
            cur = pp['pageInfo']['endCursor']
        missing = [p for p in _winner_products(tok) if p not in have]
        if not missing: return 0, None
        if dry:
            print(f"  would add {len(missing)} winner(s) to Best Sellers collection (DRY)")
            return 0, None
        M = ('mutation($id:ID!,$pids:[ID!]!){collectionAddProductsV2(id:$id,productIds:$pids){'
             'userErrors{field message}}}')
        added = 0
        for i in range(0, len(missing), 250):
            batch = [f"gid://shopify/Product/{p}" for p in missing[i:i + 250]]
            j = requests.post(f"https://{SHOP}/admin/api/{SHOP_API}/graphql.json",
                              headers={'X-Shopify-Access-Token': tok, 'Content-Type': 'application/json'},
                              json={'query': M, 'variables': {'id': cgid, 'pids': batch}}, timeout=60).json()
            errs = ((j.get('data') or {}).get('collectionAddProductsV2') or {}).get('userErrors') or []
            if errs: return added, f"collection add: {str(errs)[:100]}"
            added += len(batch)
        print(f"  Best Sellers collection: +{added} winner(s) added")
        return added, None
    except Exception as ex:
        return 0, f'Best Sellers sync error: {str(ex)[:120]}'

# ── WINNER PACE RULE (v11 — owner-approved 2026-07-12) ──────────────────────
# The Winners campaign has ONE kill rule, the "2.3-pace" rule:
#
#     KILL when Winners-campaign spend SINCE THE LAST SALE > that sale's revenue / 2.3
#     (winner with no sale in the lookback: allowance = product price / 2.3)
#
# Why 2.3: a kill line at 2.3-pace makes ~2.4 the survival standard, holding the
# Winners campaign blended at the owner's 2.4+ target. Every sale opens a fresh
# cycle and prepays EXACTLY its own revenue/2.3 of runway. Allowances never
# stack — past glory never pays for the present.
#
# Owner-approved design decisions (2026-07-12 session):
#   • Sales truth = Shopify, ANY channel (ads / organic / social / cross-sell).
#     Google conversions & attribution are NEVER consulted. Revenue is GROSS
#     with the same order-level discount scaling as the feed; refunds ignored.
#   • Spend = Winners campaign ONLY, per-product per-DAY (Google's finest grain).
#     The sale's own day is credited to the product — its new cycle is charged
#     from the NEXT day. Kills can only fire late (~1 day of that product's
#     spend), never early.
#   • ANCHORED cycles: judged from the REAL last sale — "if it's a real winner
#     it will hold". No fresh start at activation, no grace, no shields.
#   • Kill = DRAFT + tags draft_bad_product + l_camp + pub: stamp, and the
#     w_campaign tag is REMOVED. PERMANENT — no demotion back to Testing (a
#     demoted ex-winner's conversion history would out-compete cold imports for
#     Testing's Max-Conversions budget). Manual reactivation stays possible;
#     if it then sells, tag_new_winners re-promotes it and strips l_camp.
#   • Zero-spend products can never die (no spend, no crime): cross-sell and
#     organic sellers are immortal and only lift the blended.
#   • DORMANT until WINNER_KILL_START — before that date every run computes and
#     reports would-kills (Telegram preview) but drafts nothing.
#   • FAIL-SAFE: any error in this section skips winner kills for the run and
#     warns — it can never block or distort the testing kills. A glitchy run
#     that flags more than WINNER_KILL_CAP winners aborts (data glitch guard,
#     same philosophy as the testing KILL_CAP).
WINNER_PACE_ROAS  = 2.3                            # the ONE tunable constant
WINNER_KILL_START = datetime.date(2026, 7, 13)     # LIVE (owner 2026-07-13: "if a winner hits the rules kill it, don't wait for Jul 25")
WINNER_LOOKBACK_D = 60                             # last-sale + spend lookback window
WINNER_KILL_CAP   = 10                             # >N winner kills in one run = glitch -> abort + alert
WINNER_POOL_ALERT = 25                             # warn when the winner pool drops below this
WINNER_KILLS_LOG  = 'winner_kills_log.csv'

def _winner_products(tok):
    """Live winners (ACTIVE + w_campaign) -> {pid: {name, price}} (min variant price)."""
    Q = ('query($c:String){products(first:250,after:$c,query:"tag:w_campaign status:active"){'
         'pageInfo{hasNextPage endCursor} edges{node{legacyResourceId title '
         'priceRangeV2{minVariantPrice{amount}}}}}}')
    out = {}; cur = None
    while True:
        j = requests.post(f"https://{SHOP}/admin/api/{SHOP_API}/graphql.json",
                          headers={'X-Shopify-Access-Token': tok, 'Content-Type': 'application/json'},
                          json={'query': Q, 'variables': {'c': cur}}, timeout=60).json()
        c = j['data']['products']
        for e in c['edges']:
            n = e['node']
            out[str(n['legacyResourceId'])] = dict(
                name=n.get('title', ''),
                price=float((n.get('priceRangeV2') or {}).get('minVariantPrice', {}).get('amount') or 0))
        if c['pageInfo']['hasNextPage']: cur = c['pageInfo']['endCursor']
        else: break
    return out

def _winner_last_sales(tok, run_date):
    """pid -> {ts, date (UK), rev} of the LATEST order containing the product.
    Cancelled orders excluded; revenue = this product's lines in that order,
    scaled by the order-level discount factor (same GROSS convention as the feed)."""
    since = (run_date - datetime.timedelta(days=WINNER_LOOKBACK_D)).isoformat()
    Q = ('query($c:String){orders(first:100,after:$c,query:"created_at:>=%s -status:cancelled"){'
         'pageInfo{hasNextPage endCursor} edges{node{createdAt subtotalPriceSet{shopMoney{amount}} '
         'lineItems(first:100){edges{node{product{legacyResourceId} '
         'discountedTotalSet{shopMoney{amount}}}}}}}}}' % since)
    last = {}; cur = None; n_orders = 0
    while True:
        j = requests.post(f"https://{SHOP}/admin/api/{SHOP_API}/graphql.json",
                          headers={'X-Shopify-Access-Token': tok, 'Content-Type': 'application/json'},
                          json={'query': Q, 'variables': {'c': cur}}, timeout=90).json()
        c = j['data']['orders']
        for e in c['edges']:
            node = e['node']; n_orders += 1; ts = node['createdAt']
            d = datetime.datetime.fromisoformat(ts.replace('Z', '+00:00')).astimezone(UK).date().isoformat()
            lines = [(str(li['node']['product']['legacyResourceId']),
                      float(li['node']['discountedTotalSet']['shopMoney']['amount']))
                     for li in node['lineItems']['edges'] if li['node'].get('product')]
            line_sum = sum(a for _, a in lines)
            _sub = (node.get('subtotalPriceSet') or {}).get('shopMoney', {}).get('amount')
            factor = (float(_sub) / line_sum) if (_sub is not None and line_sum > 0) else 1.0
            per = collections.defaultdict(float)
            for pid, amt in lines: per[pid] += amt * factor
            for pid, rev in per.items():
                if pid not in last or ts > last[pid]['ts']:
                    last[pid] = dict(ts=ts, date=d, rev=rev)
        if c['pageInfo']['hasNextPage']: cur = c['pageInfo']['endCursor']
        else: break
    return last, n_orders

def _campaign_daily_spend(run_date, pids, campaign_id):
    """pid -> [(date_iso, GBP), ...] spend rows in ONE campaign, lookback window.
    Campaign-scoped by design: each tier's rule judges only that tier's spend."""
    import google_ads_connect as ga
    gt = ga.get_access_token()
    start = (run_date - datetime.timedelta(days=WINNER_LOOKBACK_D)).isoformat()
    q = (f"SELECT campaign.id, segments.date, segments.product_item_id, metrics.cost_micros "
         f"FROM shopping_performance_view "
         f"WHERE segments.date BETWEEN '{start}' AND '{run_date.isoformat()}' "
         f"AND campaign.id = {campaign_id} AND metrics.cost_micros > 0")
    out = collections.defaultdict(list)
    for row in _ads_search(ga, gt, q):
        # Google omits productItemId on non-shopping PMax rows (e.g. a £0.003
        # Champions row 2026-07-21) — a bare ['productItemId'] KeyError'd here and
        # silently skipped the champion demotion check for every run after it.
        item = row.get('segments', {}).get('productItemId')
        if not item:
            continue
        parts = str(item).lower().split('_')
        pid = parts[2] if len(parts) >= 3 and parts[0] == 'shopify' else None
        if pid and pid in pids:
            out[pid].append((row['segments']['date'], int(row['metrics'].get('costMicros', 0)) / 1e6))
    return out

def _winners_daily_spend(run_date, pids):
    return _campaign_daily_spend(run_date, pids, WINNERS_CAMPAIGN_ID)

def shopify_winner_kill(tok, pid):
    """DRAFT + add draft_bad_product / l_camp / pub: stamp + REMOVE w_campaign.
    Same publishedAt-preservation trick as shopify_draft (draft wipes publishedAt)."""
    gid = f"gid://shopify/Product/{pid}"
    pj = requests.post(f"https://{SHOP}/admin/api/{SHOP_API}/graphql.json",
                       headers={'X-Shopify-Access-Token': tok, 'Content-Type': 'application/json'},
                       json={'query': '{product(id:"%s"){publishedAt}}' % gid}, timeout=30).json()
    pub = ((pj.get('data') or {}).get('product') or {}).get('publishedAt')
    tags = ['draft_bad_product', LOST_TAG] + (['pub:' + str(pub)[:10]] if pub else [])
    M = '''mutation($id:ID!,$tags:[String!]!,$rm:[String!]!){
      productUpdate(input:{id:$id,status:DRAFT}){userErrors{message}}
      tagsAdd(id:$id,tags:$tags){userErrors{message}}
      tagsRemove(id:$id,tags:$rm){userErrors{message}} }'''
    j = requests.post(f"https://{SHOP}/admin/api/{SHOP_API}/graphql.json",
                      headers={'X-Shopify-Access-Token': tok, 'Content-Type': 'application/json'},
                      json={'query': M, 'variables': {'id': gid, 'tags': tags, 'rm': [WINNER_TAG]}},
                      timeout=30).json()
    if j.get('errors'): return str(j['errors'])[:120]
    d = j.get('data') or {}
    errs = sum([((d.get(k) or {}).get('userErrors') or []) for k in ('productUpdate', 'tagsAdd', 'tagsRemove')], [])
    return 'ok' if not errs else str(errs)[:120]

def _write_winner_kills_log(rows, run_date):
    import csv as _csv, os as _os
    new = not _os.path.exists(WINNER_KILLS_LOG)
    with open(WINNER_KILLS_LOG, 'a', newline='', encoding='utf-8') as f:
        w = _csv.writer(f)
        if new: w.writerow(['timestamp', 'data_date', 'product_id', 'name', 'cycle_opened_by',
                            'allowance', 'spent', 'outcome'])
        ts = datetime.datetime.now(UK).strftime('%Y-%m-%d %H:%M:%S')
        for r in rows:
            w.writerow([ts, run_date.isoformat(), r['pid'], r['name'], r['opened'],
                        round(r['allow'], 2), round(r['spent'], 2), r.get('outcome', '')])

def winner_pace_run(run_date, dry):
    """Evaluate every winner against the pace rule. Kills only when live (>= start
    date and not --dry). Returns dict(live, evaluated, flagged, killed, pool, err, closest)."""
    live = (run_date >= WINNER_KILL_START) and not dry
    res = dict(live=live, evaluated=0, flagged=[], killed=0, pool=0, err=None, closest=[])
    try:
        tok = shopify_token()
        winners = _winner_products(tok)
        # CHAMPIONS EXEMPT (2026-07-20): champions keep w_campaign but serve in the Champions
        # campaign — their winners-campaign spend is residual, and their last sale can be stale,
        # so judging them here would false-kill. They have their OWN trailing-2.0 demotion rule.
        champs = _champion_pids(tok)
        winners = {pid: m for pid, m in winners.items() if pid not in champs}
        res['evaluated'] = len(winners); res['pool'] = len(winners)
        if not winners: return res
        sales, n_orders = _winner_last_sales(tok, run_date)
        if n_orders == 0:   # glitch guard: a live store ALWAYS has orders in 60d
            res['err'] = f'orders pull returned 0 orders in {WINNER_LOOKBACK_D}d — glitch; winner kills skipped'
            return res
        spend = _winners_daily_spend(run_date, set(winners))
        rows = []
        for pid, m in winners.items():
            ls = sales.get(pid)
            if ls:
                allow = ls['rev'] / WINNER_PACE_ROAS
                spent = sum(v for d, v in spend.get(pid, ()) if d > ls['date'])   # charged from the day AFTER the sale
                opened = f"sale £{ls['rev']:.2f} on {ls['date']}"
            else:   # no sale in the lookback (stale) — allowance from price, spend from whole window
                allow = m['price'] / WINNER_PACE_ROAS
                spent = sum(v for _, v in spend.get(pid, ()))
                opened = f"no sale in {WINNER_LOOKBACK_D}d (price £{m['price']:.2f})"
            rows.append(dict(pid=pid, name=m['name'], allow=allow, spent=spent, opened=opened,
                             pct=(spent / allow * 100) if allow > 0 else 0.0))
        rows.sort(key=lambda x: -x['pct'])
        res['closest'] = [r for r in rows if 60 <= r['pct'] <= 100][:5]
        flagged = [r for r in rows if r['spent'] > r['allow']]
        res['flagged'] = flagged
        if not flagged: return res
        if len(flagged) > WINNER_KILL_CAP:
            res['err'] = (f'SAFETY STOP: {len(flagged)} winner kills > cap {WINNER_KILL_CAP} — '
                          f'looks like a data glitch; NOTHING drafted, investigate')
            res['flagged'] = []          # do not act, do not spam details
            return res
        for r in flagged:
            if live:
                r['outcome'] = shopify_winner_kill(tok, r['pid'])
                print(f"  winner KILL {r['pid']} -> {r['outcome']} | spent £{r['spent']:.2f} > "
                      f"allowance £{r['allow']:.2f} ({r['opened']}) | {r['name'][:40]}")
            else:
                r['outcome'] = 'DRY' if dry else 'PREVIEW (dormant)'
                print(f"  winner would-KILL {r['pid']} | spent £{r['spent']:.2f} > "
                      f"allowance £{r['allow']:.2f} ({r['opened']}) | {r['name'][:40]}")
        if live:
            res['killed'] = sum(1 for r in flagged if r.get('outcome') == 'ok')
            res['pool'] = len(winners) - res['killed']
            _write_winner_kills_log(flagged, run_date)
        return res
    except Exception as ex:
        res['err'] = f'winner rule error (skipped this run; testing kills unaffected): {str(ex)[:150]}'
        return res

# ── CHAMPIONS TIER (owner-approved 2026-07-20) ──────────────────────────────
# Third campaign for PROVEN repeat-sellers: PMax | Champions | UK (MCV, tROAS 2.0, £50/day).
# Only repeat-sellers ever absorbed extra spend productively (Jul-16 analysis) — this tier
# gives them a looser target so Google buys volume, watched by a per-product trailing floor.
#
#   ENTRY      3 lifetime orders (order count, not units; cancelled excluded)
#              -> tag c_champion (w_campaign KEPT), feed label -> c_champion,
#                 item-ids: Champions AG include + Winners AG include-nodes removed.
#   DEMOTE     Champions-campaign spend since the 3rd-last sale
#                > (revenue of the last TWO sales) / 2.0
#              = the trailing-ROAS-2.0 floor smoothed over two sale-gaps (owner-corrected
#              2026-07-20: counting 3 sales' revenue would double-count the anchor sale and
#              only enforce ~1.33). Spend charged from the day AFTER the anchor sale — late,
#              never early. Demotion is a SOFT landing: back to Winners, NOT drafted; the
#              winners pace clock restarts anchored on its own last sale.
#   RE-ENTER   2 NEW sales (dates strictly after the champ_demoted: stamp) while in Winners.
#   2nd FAIL   nothing special here — a demoted champion is a normal winner again; if it
#              breaches the winners 2.3 pace, the existing winner rule drafts it (permanent).
#   GUARDS     promotions capped (glitch), demotions capped (glitch), zero-orders pull skips
#              the whole section; failures WARN and never touch the testing/winner kills.
CHAMPION_TAG            = 'c_champion'
CHAMPION_DEMOTED_PREFIX = 'champ_demoted:'          # champ_demoted:YYYY-MM-DD, set on demotion
CHAMPIONS_CAMPAIGN_ID   = '24047674442'             # PMax | Champions | UK  (created 2026-07-20)
CHAMPIONS_AG_ID         = '6731971798'              # its asset group (listing tree mirrors Winners)
CHAMPION_ENTRY_ORDERS   = 3
CHAMPION_PACE_ROAS      = 2.0
CHAMPION_REPROMOTE_SALES = 2
CHAMPION_PROMOTE_CAP    = 25                        # >N promotions in one run = glitch -> abort section
CHAMPION_DEMOTE_CAP     = 10                        # >N demotions in one run = glitch -> abort section
CHAMPION_LOG            = 'champion_moves_log.csv'
LIFETIME_SINCE          = '2026-01-01'              # predates the store — lifetime = complete

def _champion_pids(tok):
    """ACTIVE products tagged c_champion -> set of pids."""
    Q = ('query($c:String){products(first:250,after:$c,query:"tag:c_champion status:active"){'
         'pageInfo{hasNextPage endCursor} edges{node{legacyResourceId}}}}')
    out = set(); cur = None
    while True:
        j = requests.post(f"https://{SHOP}/admin/api/{SHOP_API}/graphql.json",
                          headers={'X-Shopify-Access-Token': tok, 'Content-Type': 'application/json'},
                          json={'query': Q, 'variables': {'c': cur}}, timeout=60).json()
        c = j['data']['products']
        for e in c['edges']: out.add(str(e['node']['legacyResourceId']))
        if c['pageInfo']['hasNextPage']: cur = c['pageInfo']['endCursor']
        else: break
    return out

def _lifetime_sales(tok):
    """pid -> chronological [{ts, date(UK), rev}, ...] — ONE entry per ORDER containing the
    product, all orders lifetime. Cancelled excluded; per-order product revenue with the same
    order-level discount scaling as the feed (GROSS, refunds ignored). ~10 pages / run."""
    Q = ('query($c:String){orders(first:100,after:$c,query:"created_at:>=%s -status:cancelled"){'
         'pageInfo{hasNextPage endCursor} edges{node{createdAt subtotalPriceSet{shopMoney{amount}} '
         'lineItems(first:100){edges{node{product{legacyResourceId} '
         'discountedTotalSet{shopMoney{amount}}}}}}}}}' % LIFETIME_SINCE)
    sales = collections.defaultdict(list); cur = None; n_orders = 0
    while True:
        j = requests.post(f"https://{SHOP}/admin/api/{SHOP_API}/graphql.json",
                          headers={'X-Shopify-Access-Token': tok, 'Content-Type': 'application/json'},
                          json={'query': Q, 'variables': {'c': cur}}, timeout=90).json()
        c = j['data']['orders']
        for e in c['edges']:
            node = e['node']; n_orders += 1; ts = node['createdAt']
            d = datetime.datetime.fromisoformat(ts.replace('Z', '+00:00')).astimezone(UK).date().isoformat()
            lines = [(str(li['node']['product']['legacyResourceId']),
                      float(li['node']['discountedTotalSet']['shopMoney']['amount']))
                     for li in node['lineItems']['edges'] if li['node'].get('product')]
            line_sum = sum(a for _, a in lines)
            _sub = (node.get('subtotalPriceSet') or {}).get('shopMoney', {}).get('amount')
            factor = (float(_sub) / line_sum) if (_sub is not None and line_sum > 0) else 1.0
            per = collections.defaultdict(float)
            for pid, amt in lines: per[pid] += amt * factor
            for pid, rev in per.items():
                sales[pid].append(dict(ts=ts, date=d, rev=rev))
        if c['pageInfo']['hasNextPage']: cur = c['pageInfo']['endCursor']
        else: break
    for pid in sales: sales[pid].sort(key=lambda s: s['ts'])
    return sales, n_orders

def _demoted_date(tags):
    """Latest champ_demoted:YYYY-MM-DD stamp, or None."""
    ds = [t[len(CHAMPION_DEMOTED_PREFIX):] for t in tags if str(t).startswith(CHAMPION_DEMOTED_PREFIX)]
    return max(ds) if ds else None

def _ag_listing_state(ga, gt, ag_ids):
    """agid -> (item_subdiv_resource_name, {item_id_lower: node_resource_name}).
    Same discovery as ads_fast_path: the item-id subdivision is the SUBDIVISION whose
    case is attr1-with-no-value."""
    rows = _ads_search(ga, gt,
        "SELECT asset_group.id, asset_group_listing_group_filter.resource_name, "
        "asset_group_listing_group_filter.type, asset_group_listing_group_filter.parent_listing_group_filter, "
        "asset_group_listing_group_filter.case_value.product_custom_attribute.index, "
        "asset_group_listing_group_filter.case_value.product_custom_attribute.value, "
        "asset_group_listing_group_filter.case_value.product_item_id.value "
        f"FROM asset_group_listing_group_filter WHERE asset_group.id IN ({','.join(ag_ids)})")
    info = {}
    for agid in ag_ids:
        ag = [x for x in rows if str(x['assetGroup']['id']) == agid]
        subdiv = None; have = {}
        for x in ag:
            f = x['assetGroupListingGroupFilter']; cv = f.get('caseValue', {})
            pca = cv.get('productCustomAttribute')
            if f['type'] == 'SUBDIVISION' and pca is not None and 'value' not in pca:
                subdiv = f['resourceName']
        for x in ag:
            f = x['assetGroupListingGroupFilter']; cv = f.get('caseValue', {})
            if f.get('parentListingGroupFilter') == subdiv and cv.get('productItemId', {}).get('value'):
                have[cv['productItemId']['value'].lower()] = f['resourceName']
        if not subdiv: raise RuntimeError(f"item-id subdivision not found in AG {agid}")
        info[agid] = (subdiv, have)
    return info

def champion_ads_move(roster_pids, demote_pids, stok):
    """Listing-tree RECONCILIATION, every run (label path = backup, as with winners):
       - every roster champion's item-ids: Champions AG include + Winners include-nodes removed
         + TESTING AG item-BLOCK enforced (label-era winners were only blocked from Testing by
           custom_label_1=w_campaign; the c_champion label ends that block, so without an
           item-id exclude a champion becomes Testing-eligible again — found by the
           segmentation proof 2026-07-20, 11/14 champions leaked)
       - freshly demoted: Winners AG include restored
       - STRAY champions nodes (item-ids of products no longer in the roster): removed
    Fully idempotent + self-healing: an interrupted run (e.g. the 8-min cron cancelling a
    manual run mid-promotion) leaves tags without listing moves — the next run repairs it.
    Any failure WARNs — the c_champion/w_campaign label still moves the product on the
    next feed sync, and the rules stay campaign-scoped either way."""
    if not (roster_pids or demote_pids): return 0, None
    try:
        import google_ads_connect as ga
        gt = ga.get_access_token()
        info = _ag_listing_state(ga, gt, [CHAMPIONS_AG_ID, WINNERS_AG_ID, TESTING_AG_ID])
        ch_sub, ch_have = info[CHAMPIONS_AG_ID]
        wi_sub, wi_have = info[WINNERS_AG_ID]
        te_sub, te_have = info[TESTING_AG_ID]
        ops = []
        def _node(agid, subdiv, have, iid, node_type):
            if iid not in have:
                ops.append({'create': {'assetGroup': f"customers/{ga.CUSTOMER_ID}/assetGroups/{agid}",
                                       'type': node_type, 'listingSource': 'SHOPPING',
                                       'parentListingGroupFilter': subdiv,
                                       'caseValue': {'productItemId': {'value': iid}}}})
        def _include(agid, subdiv, have, iid):
            _node(agid, subdiv, have, iid, 'UNIT_INCLUDED')
        roster_iids = set()
        for pid in roster_pids:
            for iid in _variant_item_ids(stok, pid):
                roster_iids.add(iid)
                _include(CHAMPIONS_AG_ID, ch_sub, ch_have, iid)
                _node(TESTING_AG_ID, te_sub, te_have, iid, 'UNIT_EXCLUDED')   # never back into Testing
                if iid in wi_have: ops.append({'remove': wi_have[iid]})
        for pid in demote_pids:
            for iid in _variant_item_ids(stok, pid):
                _include(WINNERS_AG_ID, wi_sub, wi_have, iid)
        for iid, rn in ch_have.items():                     # stray cleanup (covers demotions too)
            if iid not in roster_iids: ops.append({'remove': rn})
        if ops:
            r = requests.post(f"{ga.ADS_BASE}/customers/{ga.CUSTOMER_ID}/assetGroupListingGroupFilters:mutate",
                              headers=ga._headers(gt), json={'operations': ops}, timeout=60).json()
            if 'error' in r: raise RuntimeError(str(r)[:300])
            print(f"  champion fast-path: {len(ops)} listing ops (roster {len(roster_pids)} / demoted {len(demote_pids)})")
        return len(ops), None
    except Exception as ex:
        print(f"  !! champion fast-path FAILED (label path moves them on next feed sync): {ex}")
        return 0, str(ex)[:150]

def _write_champion_log(rows):
    import csv as _csv, os as _os
    new = not _os.path.exists(CHAMPION_LOG)
    with open(CHAMPION_LOG, 'a', newline='', encoding='utf-8') as f:
        w = _csv.writer(f)
        if new: w.writerow(['timestamp', 'run_date', 'action', 'product_id', 'name',
                            'lifetime_orders', 'allowance', 'spent', 'outcome'])
        for r in rows: w.writerow(r)

def champion_run(feed, run_date, dry):
    """Promotions (3 lifetime orders / 2 fresh post-demotion) + demotions (trailing 2.0).
    Returns dict(roster, promoted, demoted, flagged, watch, err). Fail-safe: any error
    skips the section and warns — testing/winner kills are never affected."""
    res = dict(roster=0, promoted=[], demoted=[], flagged=[], watch=[], err=None)
    try:
        tok = shopify_token()
        active = {p['pid']: p for p in feed}
        champs = {pid: p for pid, p in active.items() if CHAMPION_TAG in p['tags']}
        res['roster'] = len(champs)
        sales, n_orders = _lifetime_sales(tok)
        if n_orders == 0:
            res['err'] = 'lifetime orders pull returned 0 — glitch; champion moves skipped'
            return res

        # ---- PROMOTIONS: winners with 3+ lifetime orders (or 2 fresh ones post-demotion) ----
        cands = []
        for pid, p in active.items():
            if CHAMPION_TAG in p['tags'] or WINNER_TAG not in p['tags']: continue
            slist = sales.get(pid, [])
            if len(slist) < CHAMPION_ENTRY_ORDERS: continue
            dem = _demoted_date(p['tags'])
            if dem and len([s for s in slist if s['date'] > dem]) < CHAMPION_REPROMOTE_SALES:
                continue                       # demoted: must RE-EARN with fresh sales
            cands.append((pid, p, len(slist), dem))
        if len(cands) > CHAMPION_PROMOTE_CAP:
            res['err'] = (f'SAFETY STOP: {len(cands)} promotions > cap {CHAMPION_PROMOTE_CAP} — '
                          f'looks like a data glitch; NO champion moves this run')
            return res
        log_rows = []; ts = datetime.datetime.now(UK).strftime('%Y-%m-%d %H:%M:%S')
        for pid, p, n_life, dem in cands:
            if dry:
                out = 'DRY'
            else:
                r1 = shopify_add_tag(tok, pid, CHAMPION_TAG)
                r2 = shopify_set_label_metafield(tok, pid, CHAMPION_TAG)
                for t in [t for t in p['tags'] if str(t).startswith(CHAMPION_DEMOTED_PREFIX)]:
                    shopify_remove_tag(tok, pid, t)          # clean re-entry
                out = 'ok' if (r1 == 'ok' and r2 == 'ok') else f'{r1}/{r2}'
                p['tags'].append(CHAMPION_TAG)
            res['promoted'].append(dict(pid=pid, name=p['name'], orders=n_life,
                                        re=bool(dem), outcome=out))
            log_rows.append([ts, run_date.isoformat(), 'RE-PROMOTE' if dem else 'PROMOTE',
                             pid, p['name'], n_life, '', '', out])
            print(f"  {'would promote' if dry else 'promote'} {'(re) ' if dem else ''}champion "
                  f"{pid} ({n_life} lifetime orders) -> {out} | {p['name'][:40]}")

        # ---- DEMOTIONS: trailing floor — champion spend since 3rd-last sale vs last-2-rev/2.0 ----
        flagged = []
        judged = {pid: p for pid, p in champs.items()}        # only pre-existing champions
        if judged:
            spend = _campaign_daily_spend(run_date, set(judged), CHAMPIONS_CAMPAIGN_ID)
            for pid, p in judged.items():
                slist = sales.get(pid, [])
                if len(slist) < CHAMPION_ENTRY_ORDERS: continue      # can't compute window
                anchor = slist[-3]
                allow = (slist[-1]['rev'] + slist[-2]['rev']) / CHAMPION_PACE_ROAS
                spent = sum(v for d, v in spend.get(pid, ()) if d > anchor['date'])   # day AFTER anchor
                pct = (spent / allow * 100) if allow > 0 else 0.0
                row = dict(pid=pid, name=p['name'], allow=allow, spent=spent, pct=pct,
                           opened=f"last-2 rev £{slist[-1]['rev'] + slist[-2]['rev']:.2f}, 3rd-last {anchor['date']}")
                if spent > allow: flagged.append(row)
                elif pct >= 60: res['watch'].append(row)
            res['watch'].sort(key=lambda x: -x['pct']); res['watch'] = res['watch'][:5]
            if len(flagged) > CHAMPION_DEMOTE_CAP:
                res['err'] = (f'SAFETY STOP: {len(flagged)} demotions > cap {CHAMPION_DEMOTE_CAP} — '
                              f'looks like a data glitch; NO demotions this run')
                flagged = []
            res['flagged'] = flagged
            for r in flagged:
                pid = r['pid']; p = judged[pid]
                if dry:
                    r['outcome'] = 'DRY'
                else:
                    r1 = shopify_remove_tag(tok, pid, CHAMPION_TAG)
                    r2 = shopify_add_tag(tok, pid, f"{CHAMPION_DEMOTED_PREFIX}{run_date.isoformat()}")
                    r3 = shopify_set_label_metafield(tok, pid, WINNER_TAG)
                    r['outcome'] = 'ok' if (r1 == r2 == r3 == 'ok') else f'{r1}/{r2}/{r3}'
                    if CHAMPION_TAG in p['tags']: p['tags'].remove(CHAMPION_TAG)
                res['demoted'].append(r)
                log_rows.append([ts, run_date.isoformat(), 'DEMOTE', pid, p['name'], len(sales.get(pid, [])),
                                 round(r['allow'], 2), round(r['spent'], 2), r['outcome']])
                print(f"  {'would demote' if dry else 'demote'} champion {pid} | spent £{r['spent']:.2f} > "
                      f"allowance £{r['allow']:.2f} ({r['opened']}) | {p['name'][:40]}")

        # ---- Ads listing reconciliation: the WHOLE roster, every run (self-healing) ----
        if not dry:
            roster_now = ({pid for pid in champs} - {x['pid'] for x in res['demoted'] if x.get('outcome') == 'ok'}
                          ) | {x['pid'] for x in res['promoted'] if x['outcome'] == 'ok'}
            _, fp_err = champion_ads_move(sorted(roster_now),
                                          [x['pid'] for x in res['demoted'] if x.get('outcome') == 'ok'], tok)
            if fp_err: res['err'] = f'fast-path: {fp_err}'
        if not dry and log_rows: _write_champion_log(log_rows)
        return res
    except Exception as ex:
        res['err'] = f'champion rule error (section skipped; other kills unaffected): {str(ex)[:150]}'
        return res

def build_report(rows, outcomes, run_date, ts, n_active, n_kills, n_drafted, dry):
    wb = Workbook()
    s = wb.active; s.title = 'Summary'
    s.append(['PMax — Google Auto-Kill Run Report'])
    s.append(['Run (UK time)', ts])
    s.append(['Data date', str(run_date), run_date.strftime('%A')])
    s.append(['Mode', 'DRY-RUN (nothing drafted)' if dry else 'LIVE (products drafted)'])
    s.append(['Active products analyzed', n_active])
    s.append(['Kills found by rules', n_kills])
    s.append(['Products drafted', n_drafted])
    tiers = collections.Counter(t for (_, t, _) in rows)
    s.append(['By tier'] + [f'{k}: {v}' for k, v in sorted(tiers.items())])
    s.append([])
    s.append(['Revenue is GROSS (refunds never deducted). Cost/clicks from Google Ads; revenue from Shopify; windows UK-aligned.'])

    d = wb.create_sheet('Drafted')
    d.append(['Product ID', 'Product Name', 'Kill Tier', 'Reason', 'days_live',
              'cost_7d', 'cost_30d', 'clicks_30d', 'rev_30d', 'rev_14d', 'rev_7d', 'roas_7d', 'outcome'])
    for p, tier, why in rows:
        d.append([int(p['pid']), p['name'], tier, why, _days_live(p, run_date),
                  round(p['cost7'], 2), round(p['cost30'], 2), p['clicks30'],
                  round(p['rev30'], 2), round(p['rev14'], 2), round(p['rev7'], 2),
                  round(p['roas7'], 2), outcomes.get(p['pid'], '')])
    for c in d['A'][1:]: c.number_format = '0'

    fname = f"auto_kill_report_{run_date}_{datetime.datetime.now(UK).strftime('%H%M')}.xlsx"
    wb.save(fname)
    return fname

def send_report(subject, body, xlsx_path=None):
    if not (RESEND_API_KEY and EMAIL_TO):
        print("!! EMAIL NOT SENT — set the RESEND_API_KEY + EMAIL_TO secrets.")
        if xlsx_path: print(f"   report saved locally: {xlsx_path}")
        return False
    try:
        payload = {'from': RESEND_FROM, 'to': [EMAIL_TO], 'subject': subject, 'text': body}
        if xlsx_path:
            with open(xlsx_path, 'rb') as f:
                payload['attachments'] = [{'filename': os.path.basename(xlsx_path),
                                           'content': base64.b64encode(f.read()).decode()}]
        r = requests.post('https://api.resend.com/emails',
                          headers={'Authorization': f'Bearer {RESEND_API_KEY}'}, json=payload, timeout=30)
        if r.status_code in (200, 201):
            print(f'Email sent to {EMAIL_TO} via Resend: "{subject}"')
            return True
        print(f"!! EMAIL FAILED (Resend {r.status_code}): {r.text[:200]}")
        return False
    except Exception as ex:
        print(f"!! EMAIL FAILED: {ex}   (report saved: {xlsx_path})")
        return False

def send_telegram(text, xlsx_path=None):
    """Every-run push: a text summary + the .xlsx as a document. No daily cap on Telegram."""
    if not (TELEGRAM_TOKEN and TELEGRAM_CHAT):
        print("!! TELEGRAM NOT SENT — set TELEGRAM_BOT_TOKEN + TELEGRAM_CHAT_ID (env or _secrets_local.py).")
        return False
    base = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}"; ok = False
    try:
        r = requests.post(f"{base}/sendMessage",
                          data={'chat_id': TELEGRAM_CHAT, 'text': text, 'parse_mode': 'HTML'}, timeout=30)
        ok = (r.status_code == 200)
        if not ok: print(f"!! Telegram message failed ({r.status_code}): {r.text[:200]}")
        if xlsx_path and os.path.exists(xlsx_path):
            with open(xlsx_path, 'rb') as f:
                rd = requests.post(f"{base}/sendDocument",
                                   data={'chat_id': TELEGRAM_CHAT}, files={'document': f}, timeout=60)
            if rd.status_code != 200: print(f"!! Telegram document failed ({rd.status_code}): {rd.text[:200]}")
        if ok: print(f"Telegram sent to chat {TELEGRAM_CHAT}.")
    except Exception as ex:
        print(f"!! TELEGRAM FAILED: {ex}")
    return ok

def _kills_last_12h():
    """Rows actually DRAFTED (outcome 'ok') in the last 12h, from kills_log_auto.csv."""
    if not os.path.exists(KILLS_LOG): return []
    cutoff = datetime.datetime.now(UK) - datetime.timedelta(hours=12); out = []
    with open(KILLS_LOG, encoding='utf-8') as f:
        for row in csv.DictReader(f):
            try: t = datetime.datetime.fromisoformat(row['timestamp']).replace(tzinfo=UK)
            except Exception: continue
            if t >= cutoff and row.get('outcome') == 'ok': out.append(row)
    return out

def build_12h_report(rows, ts):
    wb = Workbook(); s = wb.active; s.title = '12h drafted'
    s.append(['Auto-Kill — 12h digest', ts]); s.append(['Products drafted (12h)', len(rows)])
    s.append(['By tier'] + [f'{k}: {v}' for k, v in sorted(collections.Counter(r['tier'] for r in rows).items())])
    s.append([]); s.append(['timestamp', 'product_id', 'name', 'tier', 'reason', 'cost_30d', 'clicks_30d', 'roas_7d'])
    for r in rows:
        s.append([r['timestamp'], int(r['product_id']), r['name'], r['tier'], r['reason'],
                  r['cost_30d'], r['clicks_30d'], r['roas_7d']])
    fname = f"auto_kill_12h_{datetime.datetime.now(UK).strftime('%Y-%m-%d_%H%M')}.xlsx"; wb.save(fname); return fname

def maybe_send_12h_email(ts, force=False):
    """Twice a day (SUMMARY_HOURS) email a TEXT digest of the last 12h kills + the .xlsx."""
    now = datetime.datetime.now(UK)
    if not (force or (now.hour in SUMMARY_HOURS and now.minute < 8)): return
    rows = _kills_last_12h()
    tiers = collections.Counter(r['tier'] for r in rows)
    lines = [f"- {r['product_id']} [{r['tier']}] {r['name'][:44]} | £{r['cost_30d']}/30d, {r['clicks_30d']} clk"
             for r in rows[:50]]
    body = (f"Auto-Kill — 12-hour digest\n{ts} (UK)\n\n"
            f"PRODUCTS DRAFTED (last 12h): {len(rows)}\n"
            f"By tier: {dict(tiers) or '-'}\n\n"
            + ("\n".join(lines) if lines else "(nothing drafted in the last 12 hours)")
            + (f"\n...and {len(rows)-50} more" if len(rows) > 50 else "")
            + "\n\n(Full list also in the attached Excel.)")
    send_report(f"12h Auto-Kill — {len(rows)} drafted", body, build_12h_report(rows, ts) if rows else None)

def _write_kills_log(rows, outcomes, run_date, ts, dry):
    new = not os.path.exists(KILLS_LOG)
    with open(KILLS_LOG, 'a', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        if new:
            w.writerow(['timestamp', 'data_date', 'mode', 'product_id', 'name', 'tier', 'reason', 'days_live',
                        'cost_7d', 'cost_30d', 'clicks_30d', 'rev_30d', 'rev_14d', 'rev_7d', 'roas_7d', 'outcome'])
        mode = 'DRY' if dry else 'LIVE'
        for p, tier, why in rows:
            w.writerow([ts, run_date, mode, p['pid'], p['name'], tier, why, _days_live(p, run_date),
                        round(p['cost7'], 2), round(p['cost30'], 2), p['clicks30'],
                        round(p['rev30'], 2), round(p['rev14'], 2), round(p['rev7'], 2),
                        round(p['roas7'], 2), outcomes.get(p['pid'], '')])

def main():
    dry = '--dry' in sys.argv or '--dry-run' in sys.argv
    run_date = datetime.datetime.now(UK).date()
    is_monday = (run_date.weekday() == 0)
    ts = datetime.datetime.now(UK).strftime('%Y-%m-%d %H:%M:%S')

    logf = open(RUN_LOG, 'a', encoding='utf-8')
    logf.write(f"\n{'#'*72}\n# AUTO RUN {ts} (UK) | GOOGLE-DIRECT | {'DRY' if dry else 'LIVE'} | data-date {run_date}\n{'#'*72}\n")
    real = sys.stdout; sys.stdout = _Tee(real, logf)
    try:
        print(f"== kill_engine GOOGLE AUTO == {run_date} ({run_date.strftime('%A')}) | "
              f"Mon tiers 3&4 {'ON' if is_monday else 'OFF'} | {'DRY-RUN' if dry else 'LIVE'} | {ts}")
        feed = build_feed(run_date)
        print(f"feed: Google Ads + live Shopify -> {len(feed)} active products")

        # GLITCH GUARD: a live store ALWAYS has some Shopify revenue over 30 days. If the orders
        # pull comes back empty (£0 across EVERY product) it's a data glitch (failed/empty Shopify
        # response), not reality — every spending product would falsely read £0-revenue and get
        # killed. So abort + alert and draft NOTHING. This caps nothing real: a legit big batch
        # still has normal order data; only the empty-data glitch trips it.
        total_rev30 = sum(p.get('rev30', 0) for p in feed)
        if total_rev30 <= 0:
            msg = (f"ABORTED — Shopify returned £0 revenue across ALL {len(feed)} active products over 30 days. "
                   f"That's an empty/failed orders pull, not real sales. NOTHING was drafted; re-run after Shopify recovers.")
            print("!! " + msg)
            send_telegram(f"⚠️ <b>Auto-Kill ABORTED</b>\n{ts} UK\n{msg}")
            send_report("ALERT: auto-kill ABORTED — no Shopify orders data",
                        f"PMax auto-kill — {ts} (UK)\nMode: {'DRY-RUN' if dry else 'LIVE'}\n\n{msg}")
            return

        # WINNERS first: tag new 1+ sale actives, then EXEMPT all tagged from the kill rules
        new_winners = tag_new_winners(feed, dry)
        exempt = sum(1 for p in feed if WINNER_TAG in p['tags'])
        print(f"winners: +{len(new_winners)} newly tagged | {exempt} total exempt from kill rules")
        fp_err = None
        if new_winners and not dry:
            _, fp_err = ads_fast_path(new_winners, shopify_token())   # instant campaign move (label = backup)

        # BEST SELLERS: every active winner belongs to the storefront collection (add-only)
        bs_added, bs_err = sync_bestseller_collection(shopify_token(), dry)

        # CHAMPIONS TIER (2026-07-20): promote 3-lifetime-order winners into the Champions
        # campaign; demote champions whose trailing 2-gap window fell below 2.0. Runs BEFORE
        # the winner pace rule so fresh promotions are already champion-exempt this same run.
        ch = champion_run(feed, run_date, dry)
        print(f"champions: roster {ch['roster']} | promoted {len(ch['promoted'])} | "
              f"demoted {len(ch['demoted'])}" + (f" | !! {ch['err']}" if ch['err'] else ""))
        for r in ch['watch']:
            print(f"  champion watch {r['pct']:.0f}%: {r['pid']} spent £{r['spent']:.2f} of "
                  f"£{r['allow']:.2f} ({r['opened']}) | {r['name'][:40]}")

        # WINNER PACE RULE (v11): judge the Winners campaign at 2.3-pace
        w = winner_pace_run(run_date, dry)
        print(f"winner pace ({WINNER_PACE_ROAS}): {w['evaluated']} evaluated | "
              f"{len(w['flagged'])} over allowance | killed {w['killed']} | "
              f"{'LIVE' if w['live'] else ('DRY' if dry else 'dormant -> live ' + WINNER_KILL_START.isoformat())}"
              + (f" | !! {w['err']}" if w['err'] else ""))
        for r in w['closest']:
            print(f"  pace watch {r['pct']:.0f}%: {r['pid']} spent £{r['spent']:.2f} of £{r['allow']:.2f} ({r['opened']}) | {r['name'][:40]}")

        kills = []
        for p in feed:
            if WINNER_TAG in p['tags']:
                continue                      # winners: separate rules later — never killed here
            dec, tier, why = evaluate(p, run_date, is_monday)
            if dec == 'KILL':
                kills.append((p, tier, why))

        # no cap — draft EVERY product the rules flag
        to_draft = kills
        print(f"kills found: {len(kills)}")
        outcomes = {}
        if to_draft:
            wtok = None if dry else shopify_token()
            for p, tier, why in to_draft:
                res = 'DRY (not drafted)' if dry else shopify_draft(wtok, p['pid'])
                outcomes[p['pid']] = res
                print(f"  {'would draft' if dry else 'draft'} {p['pid']} [{tier}] -> {res} | {p['name'][:42]}")
        drafted = 0 if dry else sum(1 for v in outcomes.values() if v == 'ok')

        xlsx = build_report(to_draft, outcomes, run_date, ts, len(feed), len(kills), drafted, dry)
        print(f"report: {xlsx}")

        # log FIRST so the 12h digest can include this run's kills
        _write_kills_log(to_draft, outcomes, run_date, ts, dry)
        print(f"logs: run -> {RUN_LOG} | kills -> {KILLS_LOG}")

        # TELEGRAM — every run: run stats + the .xlsx
        n = len(to_draft) if dry else drafted
        tg = (f"🤖 <b>Auto-Kill</b> — {n} drafted{' (DRY)' if dry else ''}\n"
              f"{ts} UK · {run_date.strftime('%a')}\n"
              f"Active: {len(feed)} | winners exempt: {exempt} | kills found: {len(kills)} | drafted: {n}")
        if new_winners:
            tg += ("\n🏆 <b>new winners → w_campaign:</b> "
                   + ", ".join(f"<code>{p['pid']}</code>" for p in new_winners[:10])
                   + (f" +{len(new_winners)-10} more" if len(new_winners) > 10 else ""))
            tg += ("\n⚡ moved to Winners campaign instantly" if not fp_err
                   else f"\n⚠️ fast-path failed ({html.escape(fp_err[:80])}) — label moves it on next feed sync")
        if to_draft:
            DETAIL = 15                                     # full reason+metrics for up to 15; rest in the Excel
            tg += "\n\n" + "\n\n".join(_fmt_kill(p, tier, why, run_date) for p, tier, why in to_draft[:DETAIL])
            if len(to_draft) > DETAIL:
                tg += f"\n\n…+{len(to_draft)-DETAIL} more — full reasons &amp; metrics in the attached Excel."

        # WINNERS pace section — one status line every run + detail per kill/preview
        tg += (f"\n\n🎯 <b>Winners pace {WINNER_PACE_ROAS}</b>: {w['evaluated']} checked, "
               f"{len(w['flagged'])} over allowance"
               + ("" if w['live'] else (" (DRY)" if dry else f" — dormant, live {WINNER_KILL_START.strftime('%d %b')}")))
        for r in w['flagged'][:8]:
            tg += (f"\n🔻 <b>{html.escape(r['name'][:42])}</b> <code>{r['pid']}</code>\n"
                   f"   spent £{r['spent']:.2f} &gt; allowance £{r['allow']:.2f} ({html.escape(r['opened'])})"
                   + (f" → {html.escape(str(r.get('outcome', '')))}" if w['live'] else " → would kill"))
        if len(w['flagged']) > 8:
            tg += f"\n   …+{len(w['flagged']) - 8} more — see winner_kills_log.csv"
        if w['err']:
            tg += f"\n⚠️ {html.escape(w['err'])}"

        # CHAMPIONS section — roster + every move, every run
        ch_now = (ch['roster'] + sum(1 for x in ch['promoted'] if x['outcome'] in ('ok', 'DRY'))
                  - sum(1 for x in ch['demoted'] if x.get('outcome') in ('ok', 'DRY')))
        tg += (f"\n\n👑 <b>Champions (tROAS 2.0, trailing pace {CHAMPION_PACE_ROAS})</b>: "
               f"roster {ch_now} | promoted {len(ch['promoted'])} | demoted {len(ch['demoted'])}")
        for x in ch['promoted'][:10]:
            tg += (f"\n⬆️ <b>{html.escape(x['name'][:42])}</b> <code>{x['pid']}</code> — "
                   f"{x['orders']} lifetime orders{' (re-promoted)' if x['re'] else ''}"
                   + (f" → {html.escape(str(x['outcome']))}" if x['outcome'] not in ('ok', 'DRY') else ''))
        for r in ch['demoted'][:10]:
            tg += (f"\n⬇️ <b>{html.escape(r['name'][:42])}</b> <code>{r['pid']}</code>\n"
                   f"   spent £{r['spent']:.2f} &gt; allowance £{r['allow']:.2f} "
                   f"({html.escape(r['opened'])}) → back to Winners")
        for r in ch['watch'][:3]:
            tg += (f"\n👀 champion watch {r['pct']:.0f}%: <code>{r['pid']}</code> "
                   f"£{r['spent']:.2f} of £{r['allow']:.2f}")
        if ch['err']:
            tg += f"\n⚠️ {html.escape(ch['err'])}"
        if bs_added:
            tg += f"\n🛍️ Best Sellers collection: +{bs_added} winner(s) added"
        if bs_err:
            tg += f"\n⚠️ {html.escape(bs_err)}"
        if w['live'] and w['killed'] and w['pool'] < WINNER_POOL_ALERT:
            tg += (f"\n⚠️ <b>WINNER POOL LOW: {w['pool']} left</b> (&lt;{WINNER_POOL_ALERT}) — "
                   f"consider cutting the Winners £100/day budget (your manual call).")
        send_telegram(tg, xlsx)

        # RESEND email — twice a day (SUMMARY_HOURS): TEXT digest of last 12h + the .xlsx
        maybe_send_12h_email(ts, force=('--test' in sys.argv or '--email' in sys.argv))
    except Exception:
        import traceback
        err = traceback.format_exc()
        print("!! AUTO-KILL RUN FAILED — nothing further drafted:\n" + err)
        try:
            send_telegram(f"❌ <b>Auto-Kill FAILED</b>\n{ts} UK\n<pre>{err[-500:]}</pre>")
            send_report(f"AUTO-KILL FAILED — {run_date}",
                        f"auto-kill run FAILED at {ts} (UK).\n\nError:\n{err}")
        except Exception:
            pass
    finally:
        sys.stdout = real
        logf.close()

if __name__ == '__main__':
    main()
